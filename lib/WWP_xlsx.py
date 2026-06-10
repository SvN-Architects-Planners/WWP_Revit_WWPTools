# -*- coding: utf-8 -*-
"""Minimal openpyxl-compatible .xlsx reader/writer for IronPython.

openpyxl is a CPython-only library and cannot import cleanly under IronPython
(Python 3 syntax it doesn't fully support), while pyRevit 6.4's CPython engine
breaks the WPF event delegates these tools rely on. So neither engine could run
the openpyxl-based scripts.

An .xlsx file is just a ZIP of XML parts, and IronPython has full access to
.NET (System.IO.Compression + System.Xml.Linq). This module reads and writes
.xlsx directly through .NET and exposes the small subset of the openpyxl API
the WWPTools scripts actually use, so callers only swap their import:

    from openpyxl import Workbook, load_workbook       ->  from WWP_xlsx import Workbook, load_workbook
    from openpyxl.utils import get_column_letter        ->  from WWP_xlsx import get_column_letter
    from openpyxl.styles import Font                     ->  from WWP_xlsx import Font

Supported surface:
  Reading  -- load_workbook(path, data_only=True); wb.active / wb.worksheets /
              wb.sheetnames / wb[name]; ws.title / ws.max_row / ws.max_column;
              ws.cell(row, column).value; ws.cell(...).fill (patternType,
              fgColor.rgb, start_color.rgb); ws.iter_rows(values_only=True).
  Writing  -- Workbook(); ws.append(row); ws.cell(row, column, value=...);
              cell.value / cell.number_format / cell.font; ws.title;
              wb.create_sheet(title, index); wb.remove(ws); ws.columns;
              ws.column_dimensions[letter].width; ws.freeze_panes; wb.save(path).

Known limitations (accepted): date cells read as raw Excel serial numbers;
appending into an existing workbook preserves cell *values* of pre-existing
sheets but not their original styling/formulas/charts; .xlsm macros are not
preserved on save.
"""

import os

try:
    import clr
except ImportError:  # Allows the pure-Python helpers to be imported under CPython for testing.
    clr = None

_NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_NS_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
_NS_CONTENT_TYPES = "http://schemas.openxmlformats.org/package/2006/content-types"

_BUILTIN_TEXT_NUMFMT_ID = 49  # "@"


def _add_refs():
    clr.AddReference("System.IO.Compression")
    clr.AddReference("System.IO.Compression.FileSystem")
    clr.AddReference("System.Xml.Linq")


# ---------------------------------------------------------------------------
# Column / cell-reference helpers (pure Python).
# ---------------------------------------------------------------------------


def get_column_letter(index):
    """1-based column index -> letters (1 -> 'A', 27 -> 'AA')."""
    if index < 1:
        index = 1
    letters = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def column_index_from_string(letters):
    """Column letters -> 1-based index ('A' -> 1, 'AA' -> 27)."""
    index = 0
    for ch in letters or "":
        if ch.isalpha():
            index = index * 26 + (ord(ch.upper()) - 64)
    return index


def _split_ref(ref):
    """'B12' -> (row=12, col=2). Missing parts default to 0."""
    letters = ""
    digits = ""
    for ch in ref or "":
        if ch.isalpha():
            letters += ch
        elif ch.isdigit():
            digits += ch
    col = column_index_from_string(letters) if letters else 0
    row = int(digits) if digits else 0
    return row, col


def _xml_escape(text):
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


# ---------------------------------------------------------------------------
# Cell / fill model.
# ---------------------------------------------------------------------------


class Font(object):
    def __init__(self, bold=False, **kwargs):
        self.bold = bool(bold)


class _Color(object):
    def __init__(self, rgb=None):
        self.rgb = rgb


class _Fill(object):
    def __init__(self, pattern_type=None, rgb=None):
        self.patternType = pattern_type
        color = _Color(rgb)
        self.fgColor = color
        self.start_color = color


class Cell(object):
    def __init__(self, worksheet, row, column, value=None):
        self._ws = worksheet
        self.row = row
        self.column = column
        self.value = value
        self.number_format = "General"
        self.font = None
        self._fill = None

    @property
    def column_letter(self):
        return get_column_letter(self.column)

    @property
    def coordinate(self):
        return "{}{}".format(self.column_letter, self.row)

    def _get_fill(self):
        return self._fill if self._fill is not None else _Fill()

    def _set_fill(self, value):
        self._fill = value

    fill = property(_get_fill, _set_fill)


class _ColumnDimension(object):
    def __init__(self):
        self.width = None


class _ColumnDimensions(object):
    def __init__(self):
        self._dims = {}

    def __getitem__(self, letter):
        dim = self._dims.get(letter)
        if dim is None:
            dim = _ColumnDimension()
            self._dims[letter] = dim
        return dim

    def items(self):
        return self._dims.items()


class Worksheet(object):
    def __init__(self, title="Sheet"):
        self.title = title
        self._cells = {}
        self._max_row = 0
        self._max_col = 0
        self.freeze_panes = None
        self.column_dimensions = _ColumnDimensions()
        # Marks sheets written this session so save() only sends those to the
        # C# service, letting ClosedXML preserve untouched sheets on append.
        self._dirty = False

    def _touch(self, row, column):
        if row > self._max_row:
            self._max_row = row
        if column > self._max_col:
            self._max_col = column

    def cell(self, row, column, value=None):
        key = (row, column)
        cell = self._cells.get(key)
        if cell is None:
            cell = Cell(self, row, column)
            self._cells[key] = cell
            self._touch(row, column)
        if value is not None:
            cell.value = value
            self._dirty = True
        return cell

    @property
    def max_row(self):
        return self._max_row

    @property
    def max_column(self):
        return self._max_col

    def append(self, values):
        row = self._max_row + 1
        self._max_row = row
        column = 0
        for value in values:
            column += 1
            self.cell(row, column).value = value
        self._dirty = True

    def iter_rows(self, values_only=False, min_row=1, max_row=None, max_col=None):
        last_row = self._max_row if max_row is None else max_row
        last_col = self._max_col if max_col is None else max_col
        for row in range(min_row, last_row + 1):
            if values_only:
                yield tuple(
                    self._cells[(row, col)].value if (row, col) in self._cells else None
                    for col in range(1, last_col + 1)
                )
            else:
                yield tuple(self.cell(row, col) for col in range(1, last_col + 1))

    @property
    def columns(self):
        result = []
        for col in range(1, self._max_col + 1):
            result.append(tuple(self.cell(row, col) for row in range(1, self._max_row + 1)))
        return result


class Workbook(object):
    def __init__(self):
        self._sheets = []
        first = Worksheet("Sheet")
        self._sheets.append(first)
        self.active = first
        # Set by load_workbook so save() can ask the C# service to append into
        # the original file (preserving untouched sheets with full fidelity).
        self._source_path = None
        self._removed = []

    @property
    def sheetnames(self):
        return [sheet.title for sheet in self._sheets]

    @property
    def worksheets(self):
        return list(self._sheets)

    def __getitem__(self, name):
        for sheet in self._sheets:
            if sheet.title == name:
                return sheet
        raise KeyError(name)

    def __contains__(self, name):
        return name in self.sheetnames

    def create_sheet(self, title=None, index=None):
        sheet = Worksheet(title or "Sheet{}".format(len(self._sheets) + 1))
        sheet._dirty = True
        if index is None:
            self._sheets.append(sheet)
        else:
            self._sheets.insert(index, sheet)
        return sheet

    def remove(self, sheet):
        if sheet in self._sheets:
            if sheet.title and sheet.title not in self._removed:
                self._removed.append(sheet.title)
            self._sheets.remove(sheet)
            if self.active is sheet:
                self.active = self._sheets[0] if self._sheets else None

    def save(self, path):
        # Plan B: prefer the compiled WWPTools.IO Excel writer (ClosedXML) for
        # full-fidelity output and append. Fall back to the native .NET writer if
        # the DLL is unavailable or the service errors, so writes never fail hard.
        try:
            _write_via_service(self, path)
            return
        except Exception:
            pass
        _save_workbook(self, path)


# ---------------------------------------------------------------------------
# Reading (.NET zip + XML).
# ---------------------------------------------------------------------------


def _read_entry_text(archive, name):
    entry = archive.GetEntry(name)
    if entry is None:
        return None
    from System.IO import StreamReader

    stream = entry.Open()
    try:
        reader = StreamReader(stream)
        try:
            return reader.ReadToEnd()
        finally:
            reader.Close()
    finally:
        stream.Close()


def _descendants(node, local):
    found = []
    if node is None:
        return found
    for element in node.Descendants():
        if element.Name.LocalName == local:
            found.append(element)
    return found


def _first(node, local):
    if node is None:
        return None
    for element in node.Descendants():
        if element.Name.LocalName == local:
            return element
    return None


def _child_elements(node, local):
    found = []
    if node is None:
        return found
    for element in node.Elements():
        if element.Name.LocalName == local:
            found.append(element)
    return found


def _attr(element, local):
    if element is None:
        return None
    for attribute in element.Attributes():
        if attribute.Name.LocalName == local:
            return attribute.Value
    return None


def _parse_shared_strings(text):
    strings = []
    if not text:
        return strings
    from System.Xml.Linq import XDocument

    doc = XDocument.Parse(text)
    for si in _descendants(doc.Root, "si"):
        parts = [t.Value or "" for t in _descendants(si, "t")]
        strings.append("".join(parts))
    return strings


def _parse_styles(text):
    """Returns (cellxf_fill_ids, fills) where fills[i] = (patternType, rgb)."""
    cellxf_fill_ids = []
    fills = []
    if not text:
        return cellxf_fill_ids, fills
    from System.Xml.Linq import XDocument

    doc = XDocument.Parse(text)
    fills_node = _first(doc.Root, "fills")
    for fill in _child_elements(fills_node, "fill"):
        pattern = _first(fill, "patternFill")
        pattern_type = _attr(pattern, "patternType")
        fg = _first(pattern, "fgColor") if pattern is not None else None
        rgb = _attr(fg, "rgb")
        fills.append((pattern_type, rgb))

    cellxfs_node = _first(doc.Root, "cellXfs")
    for xf in _child_elements(cellxfs_node, "xf"):
        try:
            cellxf_fill_ids.append(int(_attr(xf, "fillId") or 0))
        except Exception:
            cellxf_fill_ids.append(0)
    return cellxf_fill_ids, fills


def _cell_raw_value(cell):
    for child in cell.Elements():
        if child.Name.LocalName == "v":
            return child.Value
    return None


def _coerce_value(cell, shared):
    cell_type = _attr(cell, "t")
    if cell_type == "inlineStr":
        parts = [t.Value or "" for t in _descendants(cell, "t")]
        return "".join(parts)
    raw = _cell_raw_value(cell)
    if raw is None:
        return None
    if cell_type == "s":
        try:
            idx = int(raw)
            return shared[idx] if 0 <= idx < len(shared) else ""
        except Exception:
            return ""
    if cell_type == "b":
        return raw not in ("0", "", None)
    if cell_type == "str":
        return raw
    # Numeric (no type or t="n"). Dates are serial numbers with a style; this
    # reader returns the raw number rather than a datetime.
    try:
        number = float(raw)
        return int(number) if number == int(number) else number
    except Exception:
        return raw


def _fill_for_style(style_index, cellxf_fill_ids, fills):
    if style_index is None:
        return None
    try:
        idx = int(style_index)
    except Exception:
        return None
    if idx < 0 or idx >= len(cellxf_fill_ids):
        return None
    fill_id = cellxf_fill_ids[idx]
    if fill_id < 0 or fill_id >= len(fills):
        return None
    pattern_type, rgb = fills[fill_id]
    if pattern_type is None and rgb is None:
        return None
    return _Fill(pattern_type, rgb)


def _populate_sheet(worksheet, text, shared, cellxf_fill_ids, fills):
    if not text:
        return
    from System.Xml.Linq import XDocument

    doc = XDocument.Parse(text)
    for row_node in _descendants(doc.Root, "row"):
        fallback_col = 0
        for cell_node in row_node.Elements():
            if cell_node.Name.LocalName != "c":
                continue
            ref = _attr(cell_node, "r")
            if ref:
                row_idx, col_idx = _split_ref(ref)
            else:
                fallback_col += 1
                row_idx = int(_attr(row_node, "r") or 0)
                col_idx = fallback_col
            if row_idx <= 0 or col_idx <= 0:
                continue
            cell = worksheet.cell(row_idx, col_idx)
            cell.value = _coerce_value(cell_node, shared)
            fill = _fill_for_style(_attr(cell_node, "s"), cellxf_fill_ids, fills)
            if fill is not None:
                cell._fill = fill


def _resolve_sheet_targets(archive):
    """Returns (ordered list of (name, target_path), active_index)."""
    wb_text = _read_entry_text(archive, "xl/workbook.xml")
    rels_text = _read_entry_text(archive, "xl/_rels/workbook.xml.rels")
    if not wb_text:
        return [], 0
    from System.Xml.Linq import XDocument

    workbook = XDocument.Parse(wb_text)

    rel_targets = {}
    if rels_text:
        rels = XDocument.Parse(rels_text)
        for rel in _descendants(rels.Root, "Relationship"):
            rel_targets[_attr(rel, "Id")] = _attr(rel, "Target")

    active_index = 0
    for view in _descendants(workbook.Root, "workbookView"):
        try:
            active_index = int(_attr(view, "activeTab") or 0)
        except Exception:
            active_index = 0
        break

    sheets = []
    for sheet in _descendants(workbook.Root, "sheet"):
        name = _attr(sheet, "name") or "Sheet"
        rid = _attr(sheet, "id")  # r:id -> local name "id"
        target = rel_targets.get(rid)
        if target:
            if target.startswith("/"):
                target = target.lstrip("/")
            elif not target.startswith("xl/"):
                target = "xl/" + target
        sheets.append((name, target))
    return sheets, active_index


def load_workbook(path, data_only=True, read_only=False, keep_vba=False, **kwargs):
    _add_refs()
    from System.IO.Compression import ZipFile

    archive = ZipFile.OpenRead(path)
    try:
        shared = _parse_shared_strings(_read_entry_text(archive, "xl/sharedStrings.xml"))
        cellxf_fill_ids, fills = _parse_styles(_read_entry_text(archive, "xl/styles.xml"))
        sheets, active_index = _resolve_sheet_targets(archive)

        workbook = Workbook()
        workbook._sheets = []
        workbook.active = None
        if not sheets:
            sheets = [("Sheet", "xl/worksheets/sheet1.xml")]
        for position, (name, target) in enumerate(sheets):
            worksheet = Worksheet(name)
            text = _read_entry_text(archive, target) if target else None
            if text is None:
                text = _read_entry_text(archive, "xl/worksheets/sheet{}.xml".format(position + 1))
            _populate_sheet(worksheet, text, shared, cellxf_fill_ids, fills)
            worksheet._dirty = False  # loaded content is not a pending write
            workbook._sheets.append(worksheet)
        if not workbook._sheets:
            workbook._sheets.append(Worksheet("Sheet"))
        if active_index < 0 or active_index >= len(workbook._sheets):
            active_index = 0
        workbook.active = workbook._sheets[active_index]
        workbook._source_path = path
        return workbook
    finally:
        archive.Dispose()


# ---------------------------------------------------------------------------
# Writing (build XML parts, emit .NET zip).
# ---------------------------------------------------------------------------


def _freeze_pane_xml(freeze):
    row, col = _split_ref(freeze)
    if row <= 0 and col <= 0:
        return ""
    x_split = col - 1 if col > 0 else 0
    y_split = row - 1 if row > 0 else 0
    if x_split > 0 and y_split > 0:
        active_pane = "bottomRight"
    elif x_split > 0:
        active_pane = "topRight"
    else:
        active_pane = "bottomLeft"
    attrs = []
    if x_split > 0:
        attrs.append('xSplit="{}"'.format(x_split))
    if y_split > 0:
        attrs.append('ySplit="{}"'.format(y_split))
    attrs.append('topLeftCell="{}"'.format(_xml_escape(freeze)))
    attrs.append('activePane="{}"'.format(active_pane))
    attrs.append('state="frozen"')
    return (
        '<sheetViews><sheetView workbookViewId="0"><pane {} /></sheetView></sheetViews>'.format(
            " ".join(attrs)
        )
    )


def _collect_styles(workbook):
    """Assign cellXfs indices for the (bold, number_format) combos used.

    Returns (style_lookup, need_bold_font). style_lookup maps
    (bold, numfmt_id) -> xf index; index 0 is the default (no bold, General).
    """
    style_lookup = {(False, 0): 0}
    need_bold_font = False
    for sheet in workbook.worksheets:
        for cell in sheet._cells.values():
            bold = bool(cell.font) and bool(getattr(cell.font, "bold", False))
            numfmt_id = _BUILTIN_TEXT_NUMFMT_ID if cell.number_format == "@" else 0
            if bold:
                need_bold_font = True
            key = (bold, numfmt_id)
            if key not in style_lookup:
                style_lookup[key] = len(style_lookup)
    return style_lookup, need_bold_font


def _styles_xml(style_lookup, need_bold_font):
    fonts = ['<font><sz val="11"/><name val="Calibri"/></font>']
    if need_bold_font:
        fonts.append('<font><b/><sz val="11"/><name val="Calibri"/></font>')

    ordered = sorted(style_lookup.items(), key=lambda item: item[1])
    xfs = []
    for (bold, numfmt_id), _index in ordered:
        font_id = 1 if (bold and need_bold_font) else 0
        attrs = [
            'numFmtId="{}"'.format(numfmt_id),
            'fontId="{}"'.format(font_id),
            'fillId="0"',
            'borderId="0"',
            'xfId="0"',
        ]
        if numfmt_id != 0:
            attrs.append('applyNumberFormat="1"')
        if font_id != 0:
            attrs.append('applyFont="1"')
        xfs.append("<xf {} />".format(" ".join(attrs)))

    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<styleSheet xmlns="{ns}">'
        '<fonts count="{font_count}">{fonts}</fonts>'
        '<fills count="2">'
        '<fill><patternFill patternType="none"/></fill>'
        '<fill><patternFill patternType="gray125"/></fill>'
        "</fills>"
        '<borders count="1"><border/></borders>'
        '<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
        '<cellXfs count="{xf_count}">{xfs}</cellXfs>'
        "</styleSheet>"
    ).format(
        ns=_NS_MAIN,
        font_count=len(fonts),
        fonts="".join(fonts),
        xf_count=len(xfs),
        xfs="".join(xfs),
    )


def _cell_xml(cell, style_lookup):
    value = cell.value
    if value is None:
        return ""
    bold = bool(cell.font) and bool(getattr(cell.font, "bold", False))
    numfmt_id = _BUILTIN_TEXT_NUMFMT_ID if cell.number_format == "@" else 0
    style_index = style_lookup.get((bold, numfmt_id), 0)
    ref = cell.coordinate
    style_attr = ' s="{}"'.format(style_index) if style_index else ""

    if isinstance(value, bool):
        return '<c r="{}"{} t="b"><v>{}</v></c>'.format(ref, style_attr, 1 if value else 0)
    if isinstance(value, (int, float)):
        return '<c r="{}"{}><v>{}</v></c>'.format(ref, style_attr, value)
    text = _xml_escape(value if isinstance(value, str) else str(value))
    return '<c r="{}"{} t="inlineStr"><is><t xml:space="preserve">{}</t></is></c>'.format(
        ref, style_attr, text
    )


def _sheet_xml(sheet, style_lookup):
    parts = [
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
        '<worksheet xmlns="{}">'.format(_NS_MAIN),
    ]
    if sheet.freeze_panes:
        parts.append(_freeze_pane_xml(sheet.freeze_panes))

    cols = []
    for letter, dim in sheet.column_dimensions.items():
        if dim.width is None:
            continue
        col_index = column_index_from_string(letter)
        if col_index <= 0:
            continue
        cols.append(
            '<col min="{0}" max="{0}" width="{1}" customWidth="1"/>'.format(col_index, dim.width)
        )
    if cols:
        parts.append("<cols>{}</cols>".format("".join(cols)))

    parts.append("<sheetData>")
    for row in range(1, sheet._max_row + 1):
        cells_xml = []
        for col in range(1, sheet._max_col + 1):
            cell = sheet._cells.get((row, col))
            if cell is None:
                continue
            cells_xml.append(_cell_xml(cell, style_lookup))
        row_body = "".join(cells_xml)
        if row_body:
            parts.append('<row r="{}">{}</row>'.format(row, row_body))
        else:
            parts.append('<row r="{}"/>'.format(row))
    parts.append("</sheetData></worksheet>")
    return "".join(parts)


def _content_types_xml(sheet_count):
    overrides = [
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>',
        '<Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>',
    ]
    for i in range(1, sheet_count + 1):
        overrides.append(
            '<Override PartName="/xl/worksheets/sheet{}.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'.format(
                i
            )
        )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="{ns}">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        "{overrides}"
        "</Types>"
    ).format(ns=_NS_CONTENT_TYPES, overrides="".join(overrides))


def _root_rels_xml():
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="{ns}">'
        '<Relationship Id="rId1" Type="{rel}/officeDocument" Target="xl/workbook.xml"/>'
        "</Relationships>"
    ).format(ns=_NS_PKG_REL, rel=_NS_REL)


def _workbook_xml(sheets):
    sheet_tags = []
    for i, sheet in enumerate(sheets, start=1):
        sheet_tags.append(
            '<sheet name="{name}" sheetId="{sid}" r:id="rId{sid}"/>'.format(
                name=_xml_escape(sheet.title), sid=i
            )
        )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="{ns}" xmlns:r="{rel}">'
        "<sheets>{sheets}</sheets>"
        "</workbook>"
    ).format(ns=_NS_MAIN, rel=_NS_REL, sheets="".join(sheet_tags))


def _workbook_rels_xml(sheet_count):
    rels = []
    for i in range(1, sheet_count + 1):
        rels.append(
            '<Relationship Id="rId{0}" Type="{rel}/worksheet" Target="worksheets/sheet{0}.xml"/>'.format(
                i, rel=_NS_REL
            )
        )
    rels.append(
        '<Relationship Id="rId{0}" Type="{rel}/styles" Target="styles.xml"/>'.format(
            sheet_count + 1, rel=_NS_REL
        )
    )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="{ns}">{rels}</Relationships>'
    ).format(ns=_NS_PKG_REL, rels="".join(rels))


def _build_parts(workbook):
    sheets = workbook.worksheets
    if not sheets:
        sheets = [Worksheet("Sheet")]
    style_lookup, need_bold_font = _collect_styles(workbook)
    parts = [
        ("[Content_Types].xml", _content_types_xml(len(sheets))),
        ("_rels/.rels", _root_rels_xml()),
        ("xl/workbook.xml", _workbook_xml(sheets)),
        ("xl/_rels/workbook.xml.rels", _workbook_rels_xml(len(sheets))),
        ("xl/styles.xml", _styles_xml(style_lookup, need_bold_font)),
    ]
    for i, sheet in enumerate(sheets, start=1):
        parts.append(("xl/worksheets/sheet{}.xml".format(i), _sheet_xml(sheet, style_lookup)))
    return parts


def _save_workbook(workbook, path):
    _add_refs()
    import os

    from System.IO import FileStream, FileMode, StreamWriter
    from System.IO.Compression import ZipArchive, ZipArchiveMode
    from System.Text import UTF8Encoding

    encoding = UTF8Encoding(False)
    parts = _build_parts(workbook)

    directory = os.path.dirname(path)
    if directory and not os.path.isdir(directory):
        os.makedirs(directory)
    if os.path.isfile(path):
        os.remove(path)

    stream = FileStream(path, FileMode.Create)
    try:
        archive = ZipArchive(stream, ZipArchiveMode.Create)
        try:
            for name, text in parts:
                entry = archive.CreateEntry(name)
                entry_stream = entry.Open()
                try:
                    writer = StreamWriter(entry_stream, encoding)
                    try:
                        writer.Write(text)
                    finally:
                        writer.Close()
                finally:
                    entry_stream.Close()
        finally:
            archive.Dispose()
    finally:
        stream.Close()


# ---------------------------------------------------------------------------
# Plan B: route writes through the compiled, Revit-free WWPTools.IO DLL
# (ClosedXML) for full-fidelity output and append. Loaded lazily; all callers
# fall back to the native writer above if the DLL is unavailable.
# ---------------------------------------------------------------------------

_EXCEL_SERVICE = None
_PDF_SERVICE = None
_IO_LOAD_ERROR = None


def _runtime_is_core():
    """True under .NET (Core) 5+ (Revit 2025+), False on .NET Framework 4.8."""
    try:
        from System.Runtime.InteropServices import RuntimeInformation
        return ".NET Framework" not in (RuntimeInformation.FrameworkDescription or "")
    except Exception:
        return False


def _io_dll_path():
    cur_dir = os.path.dirname(os.path.abspath(__file__))
    name = "WWPTools.IO.net8.0-windows.dll" if _runtime_is_core() else "WWPTools.IO.net48.dll"
    return os.path.join(cur_dir, name)


def _is_remote_path(path):
    try:
        import ctypes
        if not path:
            return False
        if path.startswith("\\\\"):
            return True
        drive = os.path.splitdrive(path)[0]
        if not drive:
            return False
        return ctypes.windll.kernel32.GetDriveTypeW(drive + "\\") == 4  # DRIVE_REMOTE
    except Exception:
        return False


def _copy_local_io(dll_path):
    try:
        import shutil
        temp_root = os.environ.get("TEMP") or os.environ.get("TMP") or os.getcwd()
        temp_dir = os.path.join(temp_root, "WWPTools.IO")
        if not os.path.isdir(temp_dir):
            os.makedirs(temp_dir)
        local = os.path.join(temp_dir, os.path.basename(dll_path))
        if not os.path.isfile(local) or os.path.getmtime(local) < os.path.getmtime(dll_path):
            shutil.copy2(dll_path, local)
        return local
    except Exception:
        return dll_path


def _ensure_io_loaded():
    global _EXCEL_SERVICE, _PDF_SERVICE, _IO_LOAD_ERROR
    if _EXCEL_SERVICE is not None and _PDF_SERVICE is not None:
        return
    if _IO_LOAD_ERROR is not None:
        raise Exception(_IO_LOAD_ERROR)
    try:
        if clr is None:
            raise Exception("clr is unavailable (not running on .NET).")
        dll = _io_dll_path()
        if not os.path.isfile(dll):
            raise Exception("Missing WWPTools.IO DLL: {}".format(dll))
        load_path = _copy_local_io(dll) if _is_remote_path(dll) else dll
        if hasattr(clr, "AddReferenceToFileAndPath"):
            clr.AddReferenceToFileAndPath(load_path)
        else:
            clr.AddReference(load_path)
        from WWPTools.IO import ExcelService, PdfService
        _EXCEL_SERVICE = ExcelService
        _PDF_SERVICE = PdfService
    except Exception as exc:
        _IO_LOAD_ERROR = str(exc)
        raise


def _load_excel_service():
    _ensure_io_loaded()
    return _EXCEL_SERVICE


def load_pdf_service():
    """Returns the WWPTools.IO.PdfService type, or raises if the DLL is unavailable."""
    _ensure_io_loaded()
    return _PDF_SERVICE


def _coerce_spec_value(value):
    if value is None or isinstance(value, bool) or isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        return value
    try:
        return str(value)
    except Exception:
        return ""


def _sheet_to_spec(worksheet, index):
    max_row = worksheet._max_row
    max_col = worksheet._max_col
    cells = worksheet._cells
    rows = []
    bold = []
    text = []
    for r in range(1, max_row + 1):
        row_values = []
        for c in range(1, max_col + 1):
            cell = cells.get((r, c))
            if cell is None:
                row_values.append(None)
                continue
            row_values.append(_coerce_spec_value(cell.value))
            if cell.font is not None and getattr(cell.font, "bold", False):
                bold.append([r - 1, c - 1])
            if cell.number_format == "@":
                text.append([r - 1, c - 1])
        rows.append(row_values)

    col_widths = {}
    for letter, dim in worksheet.column_dimensions.items():
        if dim.width is None:
            continue
        idx = column_index_from_string(letter)
        if idx >= 1:
            col_widths[str(idx)] = dim.width

    spec = {"name": worksheet.title, "position": index, "rows": rows}
    if worksheet.freeze_panes:
        spec["freeze"] = worksheet.freeze_panes
    if bold:
        spec["bold"] = bold
    if text:
        spec["text"] = text
    if col_widths:
        spec["colWidths"] = col_widths
    return spec


def _build_write_spec(workbook, path):
    source = workbook._source_path
    append = bool(source) and os.path.isfile(source)
    sheets = [
        _sheet_to_spec(ws, index)
        for index, ws in enumerate(workbook._sheets)
        if ws._dirty
    ]
    return {
        "append": append,
        "sourcePath": source if append else None,
        "removed": list(workbook._removed),
        "removeDefaultSheet": True,
        "sheets": sheets,
    }


def _write_via_service(workbook, path):
    """Writes the workbook via WWPTools.IO.ExcelService. Raises on any failure
    (DLL missing or service error) so Workbook.save() falls back to native."""
    import json

    service = _load_excel_service()
    spec_json = json.dumps(_build_write_spec(workbook, path))
    service.WriteWorkbook(path, spec_json)
